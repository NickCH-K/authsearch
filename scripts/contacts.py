#!/usr/bin/env python3
"""
Helper for managing Contact Info.xlsx — the master output of the many-analyst
study contact-collection project.

Usage:
    python3 scripts/contacts.py init           # create the workbook with headers (idempotent)
    python3 scripts/contacts.py list           # print all rows
    python3 scripts/contacts.py studies        # print distinct study names already recorded
    python3 scripts/contacts.py add --study "..." --link "..." --organizer "..." \
            --email "..." --phone "..." --affiliation "..." [--notes "..."]

Design notes:
  * One row per ORGANIZER. A study with multiple organizers gets multiple rows.
  * If contact info can't be found but the study is a good one, pass
    --email "FAILED TO FIND".
  * `add` is idempotent on (study, organizer): re-adding updates the existing row
    instead of duplicating, so the script is safe to re-run after a restart.
"""
import argparse
import os
import sys

from openpyxl import Workbook, load_workbook

XLSX = os.path.join(os.path.dirname(__file__), "..", "Contact Info.xlsx")
XLSX = os.path.abspath(XLSX)

HEADERS = [
    "Study Name",
    "Link (DOI preferred)",
    "Organizer Name",
    "Organizer Email",
    "Organizer Phone",
    "Organizer Affiliation",
    "Notes",
]


def load():
    if not os.path.exists(XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(HEADERS)
        wb.save(XLSX)
    return load_workbook(XLSX)


def cmd_init(args):
    load()
    print(f"Initialized {XLSX}")


def cmd_list(args):
    wb = load()
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(i, " | ".join("" if c is None else str(c) for c in row))


def cmd_studies(args):
    wb = load()
    ws = wb.active
    seen = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0] not in seen:
            seen.append(row[0])
    for s in seen:
        print(s)


def cmd_add(args):
    wb = load()
    ws = wb.active
    new = [
        args.study,
        args.link or "",
        args.organizer or "",
        args.email or "",
        args.phone or "",
        args.affiliation or "",
        args.notes or "",
    ]
    # idempotent on (study, organizer)
    for row in ws.iter_rows(min_row=2):
        if (row[0].value or "") == args.study and (row[2].value or "") == (args.organizer or ""):
            for cell, val in zip(row, new):
                cell.value = val
            wb.save(XLSX)
            print(f"Updated existing row for {args.organizer!r} in {args.study!r}")
            return
    ws.append(new)
    wb.save(XLSX)
    print(f"Added {args.organizer!r} for {args.study!r}")


def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("init")
    sub.add_parser("list")
    sub.add_parser("studies")
    a = sub.add_parser("add")
    a.add_argument("--study", required=True)
    a.add_argument("--link", default="")
    a.add_argument("--organizer", default="")
    a.add_argument("--email", default="")
    a.add_argument("--phone", default="")
    a.add_argument("--affiliation", default="")
    a.add_argument("--notes", default="")
    args = p.parse_args()
    {"init": cmd_init, "list": cmd_list, "studies": cmd_studies, "add": cmd_add}[args.cmd](args)


if __name__ == "__main__":
    main()
