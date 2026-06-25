#!/usr/bin/env python3
"""
Add/populate the Abstract and ManyAnalystStyle columns in Contact Info.xlsx.

Usage:
  python3 scripts/annotate.py init                 # ensure the two columns exist
  python3 scripts/annotate.py set <match.json>     # set values from a JSON file
  python3 scripts/annotate.py status               # show which studies are filled

The JSON for `set` is a list of objects:
  [{"match": "<substring of Study Name>",
    "abstract": "...",
    "style": "..."}, ...]
Every row whose Study Name contains `match` gets the same abstract+style.
Idempotent: re-running overwrites those two cells for matched rows.
"""
import json
import os
import sys

from openpyxl import load_workbook

XLSX = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Contact Info.xlsx"))
ABS_HDR = "Abstract"
MAS_HDR = "ManyAnalystStyle"


def cols(ws):
    headers = [c.value for c in ws[1]]
    return headers


def ensure_cols(ws):
    headers = cols(ws)
    changed = False
    if ABS_HDR not in headers:
        ws.cell(row=1, column=len(headers) + 1, value=ABS_HDR)
        headers = cols(ws)
        changed = True
    if MAS_HDR not in headers:
        ws.cell(row=1, column=len(headers) + 1, value=MAS_HDR)
        changed = True
    return changed


def col_idx(ws, name):
    return cols(ws).index(name) + 1


def cmd_init():
    wb = load_workbook(XLSX)
    ws = wb.active
    ensure_cols(ws)
    wb.save(XLSX)
    print("Columns ensured:", ABS_HDR, MAS_HDR)


def cmd_set(path):
    data = json.load(open(path))
    wb = load_workbook(XLSX)
    ws = wb.active
    ensure_cols(ws)
    ai, mi = col_idx(ws, ABS_HDR), col_idx(ws, MAS_HDR)
    si = col_idx(ws, "Study Name")
    n = 0
    for entry in data:
        m = entry["match"]
        hit = 0
        for row in ws.iter_rows(min_row=2):
            sn = row[si - 1].value or ""
            if m in sn:
                if "abstract" in entry:
                    row[ai - 1].value = entry["abstract"]
                if "style" in entry:
                    row[mi - 1].value = entry["style"]
                hit += 1
        if hit == 0:
            print(f"  !! no rows matched: {m!r}")
        else:
            n += 1
            print(f"  set {hit} row(s) for: {m!r}")
    wb.save(XLSX)
    print(f"Updated {n} studies.")


def cmd_status():
    wb = load_workbook(XLSX)
    ws = wb.active
    ai, mi = col_idx(ws, ABS_HDR), col_idx(ws, MAS_HDR)
    si = col_idx(ws, "Study Name")
    seen = {}
    for row in ws.iter_rows(min_row=2):
        sn = row[si - 1].value or ""
        if not sn:
            continue
        a = bool(row[ai - 1].value)
        m = bool(row[mi - 1].value)
        seen[sn] = (a, m)
    done = sum(1 for a, m in seen.values() if a and m)
    print(f"{done}/{len(seen)} studies have both columns filled")
    for sn, (a, m) in seen.items():
        if not (a and m):
            print(f"  MISSING {'A' if not a else ''}{'M' if not m else ''}: {sn[:70]}")


if __name__ == "__main__":
    cmd = sys.argv[1]
    if cmd == "init":
        cmd_init()
    elif cmd == "set":
        cmd_set(sys.argv[2])
    elif cmd == "status":
        cmd_status()
